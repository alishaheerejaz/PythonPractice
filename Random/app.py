import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from alipdf.alipdf import pdf2text

pdf2text.convert()

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range (2, sheet.max_row + 1 ):
        cell = sheet.cell(row, 3)
        corrected = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected

    values = Reference(sheet,
              min_row = 2,
              max_row = sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


# from pathlib import Path
# path = Path()
#
# for file in path.glob('*'):
#     print(file)


# import random
#
# members = ['John', 'Marry', 'Tech']
# leader = random.choice(members)
# print(leader)


# for i in range(3):
#     print(random.randint(10,20))


# from ecommerce import shipping
# shipping.calc_shipping()


# import utils
#
# numbers = [5,8,3,8,0,3,6,2]
# print(utils.find_max(numbers))

# import converters
# from converters import kg_to_lbs
#
# print(kg_to_lbs(20))
# print(converters.kg_to_lbs(10))

# is_hot = False
# is_cold = False
# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("Its cold day")
#     print("wear warm clothes")
# else:
#     print("Its lovely day")
#
#
# print("Enjoy your day")
