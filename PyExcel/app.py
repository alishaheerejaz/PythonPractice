

import openpyxl

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet"]

cell = sheet["a1"]
print(sheet["a:c"])
print(sheet["a1:c3"])


sheet.append([1, 2, 3])
wb.save("transact2.xlsx")
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)


# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# print(sheet.cell(row=1, column=1))
